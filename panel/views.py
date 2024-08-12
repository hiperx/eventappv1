from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import render, redirect
from events.models import ActiveEvent
from bal.models import Bal
from piknik.models import Piknik
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse

from piknik.models import Przystanek
from django.contrib import messages
from django.http import JsonResponse
from django.db import models
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io import BytesIO
import base64
from matplotlib.dates import date2num
from datetime import datetime, timedelta
from matplotlib.dates import DateFormatter
import matplotlib.dates as mdates
from django.utils import timezone

from django.shortcuts import get_object_or_404

def uruchom_event(request):
    if request.method == 'POST':
        event_id = request.POST.get('event_id')
        if event_id:
            try:
                event = ActiveEvent.objects.get(pk=event_id)
                # Sprawdź, czy istnieją rekordy dla danego eventu w tabelach Bal i Piknik
                bal_records = Bal.objects.filter(event=event)
                piknik_records = Piknik.objects.filter(event=event)

                if bal_records.exists() or piknik_records.exists():
                    # Jeżeli istnieją rekordy dla danego eventu w jednej z tabel
                    if bal_records.exists():
                        last_modification_date = bal_records.latest('data_modyfikacji').data_modyfikacji
                    elif piknik_records.exists():
                        last_modification_date = piknik_records.latest('data_modyfikacji').data_modyfikacji
                    
                    # Sprawdź datę ostatniej modyfikacji
                    if timezone.now() - last_modification_date > timezone.timedelta(days=3):
                        # Jeżeli data ostatniej modyfikacji jest większa niż 3 dni temu, wyświetl komunikat
                        messages.error(request, "Event był już aktywny i nie można go uruchomić ponownie.")
                    else:
                        # Jeżeli data ostatniej modyfikacji jest mniejsza niż 3 dni temu, ustaw pole aktywny na True
                        event.aktywny = True
                        event.save()
                else:
                    # Jeżeli nie ma rekordów dla danego eventu w żadnej z tabel, ustaw pole aktywny na True
                    event.aktywny = True
                    event.save()

            except ActiveEvent.DoesNotExist:
                pass  # Obsłuż sytuację, gdy event o danym ID nie istnieje lub nie został znaleziony

        return redirect('panel:panel-index')  # Przekieruj gdziekolwiek po wykonaniu operacji

    events = ActiveEvent.objects.all()  # Pobierz wszystkie dostępne eventy
    return render(request, 'panel/uruchom_event.html', {'events': events})



def lista_przystankow(request):
    if request.method == 'POST':
        # Obsługa dodawania nowego przystanku
        nazwa = request.POST.get('nazwa')
        
        # Pobierz aktywne wydarzenie
        active_event = get_object_or_404(ActiveEvent, aktywny=True)
        
        # Sprawdź, czy istnieje aktywne wydarzenie
        if active_event:
            oddzial = active_event.oddzial
            Przystanek.objects.create(nazwa=nazwa, oddzial=oddzial)
        else:
            # Jeśli nie ma aktywnego wydarzenia, utwórz przystanek bez przypisanego oddziału
            Przystanek.objects.create(nazwa=nazwa)
        
        return redirect('panel:lista_przystankow')

    # Pobierz posortowaną listę przystanków
    przystanki = Przystanek.objects.all().order_by('nazwa')
    
    # Pobierz aktywne wydarzenie
    active_event = ActiveEvent.objects.filter(aktywny=True).first()
    
    return render(request, 'panel/lista_przystankow.html', {'przystanki': przystanki, 'active_event': active_event})


def edytuj_przystanek(request, przystanek_id):
    przystanek = Przystanek.objects.get(id=przystanek_id)
    if request.method == 'POST':
        # Obsługa edycji przystanku
        nazwa = request.POST.get('nazwa')
        przystanek.nazwa = nazwa
        przystanek.save()
        return redirect('panel:lista_przystankow')
    
    # Pobierz aktywne wydarzenie
    active_event = ActiveEvent.objects.filter(aktywny=True).first()    

    return render(request, 'panel/edytuj_przystanek.html', {'przystanek': przystanek, 'active_event': active_event})


def usun_przystanek(request, przystanek_id):
    przystanek = Przystanek.objects.get(id=przystanek_id)
    przystanek.delete()
    return redirect('panel:lista_przystankow')


def generate_plot_data(event_type):
    # Pobierz aktywne wydarzenie
    active_event = ActiveEvent.objects.filter(aktywny=True).first()

    # Pobierz dane z bazy w zależności od rodzaju wydarzenia (Bal lub Piknik)
    if event_type == 'bal':
        queryset = Bal.objects.filter(event=active_event, is_registred=True)
    elif event_type == 'piknik':
        queryset = Piknik.objects.filter(event=active_event, is_registred=True)
    else:
        return None

    # Grupuj dane według daty utworzenia
    data = queryset.extra({'date_created': 'date(data_utworzenia)'}).values('date_created').annotate(total_registered=models.Count('id'))

    # Przygotuj dane do wykresu
    dates = [entry['date_created'] for entry in data]
    counts = [entry['total_registered'] for entry in data]

    return dates, counts

def generate_daily_plot():
    
    # Utwórz wykres
    fig, ax = plt.subplots()

    # Pobierz dane dla Balu
    bal_dates, bal_counts = generate_plot_data('bal')
    if bal_dates and bal_counts:  # Sprawdź, czy sekwencje nie są puste
        ax.plot(bal_dates, bal_counts, label=None, marker='o')  # label='Bal',

    # Dodaj etykiety dla punktów na osi Y
    for i, txt in enumerate(bal_counts):
        ax.annotate(txt, (bal_dates[i], bal_counts[i]), textcoords="offset points", xytext=(0, 10), ha='center')

    # Pobierz dane dla Pikniku
    piknik_dates, piknik_counts = generate_plot_data('piknik')
    if piknik_dates and piknik_counts:  # Sprawdź, czy sekwencje nie są puste
        ax.plot(piknik_dates, piknik_counts, label=None, marker='o')  # label='Piknik'

    # Dodaj etykiety dla punktów na osi Y
    for i, txt in enumerate(piknik_counts):
        ax.annotate(txt, (piknik_dates[i], piknik_counts[i]), textcoords="offset points", xytext=(0, 10), ha='center')

    # Dodaj etykiety
    ax.set_xlabel('Data Zapisania')
    ax.set_ylabel('Ilość Zapisanych')

    # Dostosuj zakres osi y dodając margines, jeśli dane są niepuste
    max_y_value = max(max(bal_counts, default=0), max(piknik_counts, default=0)) + 20 #dodanie do osi y 20 glosow aby lepiej czytelniej bylo na wykresie.
    ax.set_ylim(0, max_y_value)
    
    # Ukryj legendę
    #ax.legend()
    ax.legend().set_visible(False)

#    # Daty w pionie
#    plt.xticks(rotation=90)
#    # Rozszerzenie osi x
    fig.autofmt_xdate(rotation=45)
#    # Skrócone daty
#    date_format = mdates.DateFormatter('%d.%m')
#    ax.xaxis.set_major_formatter(date_format)

    # Zapisz wykres do pliku
    image_stream = BytesIO()
    plt.savefig(image_stream, format='png')
    image_stream.seek(0)

    # Zamień obraz na dane base64
    plot_base64 = base64.b64encode(image_stream.read()).decode('utf-8')

    return plot_base64

@method_decorator(login_required, name='dispatch')

class PanelIndexView(View):
    template_name = 'panel/index.html'

    def get(self, request, *args, **kwargs):
        # Pobierz aktywne wydarzenie
        active_event = ActiveEvent.objects.filter(aktywny=True).first()
        success_message = request.GET.get('success_message', None)
        total_registered = 0  # Zainicjowana wartość domyślna
        # Sprawdź, czy aktywny event jest Bal czy Piknik
        if active_event and active_event.rodzaj_eventu == 'bal':
            # Pobierz dane z modelu Bal
            
            total_registered = Bal.objects.filter(event=active_event, is_registred=True).count()
            
            # Przetwórz dane do formatu, który będzie używany w szablonie
            
        elif active_event and active_event.rodzaj_eventu == 'piknik':
            # Pobierz dane z modelu Piknik
            total_registered = Piknik.objects.filter(event=active_event, is_registred=True).count()

        # Generuj wykres
        plot_base64 = generate_daily_plot()

        # Przekaz dane do szablonu
        context = {'active_event': active_event, 'plot_base64': plot_base64, 'success_message': success_message, 'total_registered': total_registered}
        return render(request, self.template_name, context)




class ListaOsobWypisanychView(View):
    template_name = 'panel/lista_osob_wypisanych.html'
    
    def get(self, request, *args, **kwargs):
        # Pobierz aktywne wydarzenie
        active_event = ActiveEvent.objects.filter(aktywny=True).first()
        user = request.user
        
        # Inicjalizuj dane jako puste
        data = []
        total_registered = 0  # Zainicjowana wartość domyślna
        # Sprawdź, czy aktywny event jest Bal czy Piknik
        if active_event and active_event.rodzaj_eventu == 'bal':
            # Pobierz dane z modelu Bal
            bal_data = Bal.objects.filter(event=active_event, is_registred=False)
            total_registered = Bal.objects.filter(event=active_event, is_registred=False).count()
            
            # Przetwórz dane do formatu, który będzie używany w szablonie
            data = [{'lp': idx + 1, 'login': bal.login, 'imie': bal.imie, 'nazwisko': bal.nazwisko, 'data_utworzenia': bal.data_utworzenia, 'data_modyfikacji': bal.data_modyfikacji} for idx, bal in enumerate(bal_data)]
        elif active_event and active_event.rodzaj_eventu == 'piknik':
            # Pobierz dane z modelu Piknik
            piknik_data = Piknik.objects.filter(event=active_event, is_registred=False)
            
            # Przetwórz dane do formatu, który będzie używany w szablonie
            data = [{'lp': idx + 1, 'identyfikator': piknik.identyfikator, 'login': piknik.login, 'imie': piknik.imie, 'nazwisko': piknik.nazwisko,
                     'osoba_towarzyszaca': piknik.osoba_towarzyszaca, 'liczba_dzieci': piknik.liczba_dzieci, 'wiek_dziecka_1': piknik.wiek_dziecka_1, 'wiek_dziecka_2': piknik.wiek_dziecka_2, 'wiek_dziecka_3': piknik.wiek_dziecka_3, 'wiek_dziecka_4': piknik.wiek_dziecka_4, 'wiek_dziecka_5': piknik.wiek_dziecka_5, 'wiek_dziecka_6': piknik.wiek_dziecka_6, 'wiek_dziecka_7': piknik.wiek_dziecka_7,
                     'transport_wlasny': piknik.transport_wlasny, 'przystanek': piknik.przystanek, 'data_utworzenia': piknik.data_utworzenia, 'data_modyfikacji': piknik.data_modyfikacji, 'zaakceptowane_regulamin': piknik.zaakceptowane_regulamin} for idx, piknik in enumerate(piknik_data)]
            total_registered = Piknik.objects.filter(event=active_event, is_registred=False).count()
        
        # Przekaz dane do szablonu
        context = {'data': data, 'active_event': active_event, 'total_registered': total_registered, 'user': user}
        return render(request, self.template_name, context)
    
class ListaOsobZapisanychView(View):
    template_name = 'panel/lista_osob_zapisanych.html'

    def get(self, request, *args, **kwargs):
        # Pobierz aktywne wydarzenie
        active_event = ActiveEvent.objects.filter(aktywny=True).first()
        success_message = request.GET.get('success_message', None)

        # Inicjalizuj dane jako puste
        data = []
        total_registered = 0  # Zainicjowana wartość domyślna
        # Sprawdź, czy aktywny event jest Bal czy Piknik
        if active_event and active_event.rodzaj_eventu == 'bal':
            # Pobierz dane z modelu Bal
            bal_data = Bal.objects.filter(event=active_event, is_registred=True)
            total_registered = Bal.objects.filter(event=active_event, is_registred=True).count()
            
            # Przetwórz dane do formatu, który będzie używany w szablonie
            data = [{'lp': idx + 1, 'login': bal.login, 'imie': bal.imie, 'nazwisko': bal.nazwisko, 'data_utworzenia': bal.data_utworzenia, 'data_modyfikacji': bal.data_modyfikacji} for idx, bal in enumerate(bal_data)]
        elif active_event and active_event.rodzaj_eventu == 'piknik':
            # Pobierz dane z modelu Piknik
            piknik_data = Piknik.objects.filter(event=active_event, is_registred=True)
            
            # Przetwórz dane do formatu, który będzie używany w szablonie
            data = [{'lp': idx + 1, 'identyfikator': piknik.identyfikator, 'login': piknik.login, 'imie': piknik.imie, 'nazwisko': piknik.nazwisko,
                     'osoba_towarzyszaca': piknik.osoba_towarzyszaca, 'liczba_dzieci': piknik.liczba_dzieci, 'wiek_dziecka_1': piknik.wiek_dziecka_1, 'wiek_dziecka_2': piknik.wiek_dziecka_2, 'wiek_dziecka_3': piknik.wiek_dziecka_3, 'wiek_dziecka_4': piknik.wiek_dziecka_4, 'wiek_dziecka_5': piknik.wiek_dziecka_5, 'wiek_dziecka_6': piknik.wiek_dziecka_6, 'wiek_dziecka_7': piknik.wiek_dziecka_7,
                     'transport_wlasny': piknik.transport_wlasny, 'przystanek': piknik.przystanek, 'data_utworzenia': piknik.data_utworzenia, 'data_modyfikacji': piknik.data_modyfikacji, 'zaakceptowane_regulamin': piknik.zaakceptowane_regulamin} for idx, piknik in enumerate(piknik_data)]
            total_registered = Piknik.objects.filter(event=active_event, is_registred=True).count()
        
        # Przekaz dane do szablonu
        context = {'data': data, 'active_event': active_event, 'total_registered': total_registered, 'success_message': success_message}
        return render(request, self.template_name, context)
    

class WyłączZapisyView(View):
    def get(self, request, *args, **kwargs):
        # Znajdź aktywny event
        active_event = ActiveEvent.objects.filter(aktywny=True).first()

        if active_event:
            if not active_event.aktywny:
                # Jeżeli event jest już wyłączony, zwróć odpowiedź JSON z komunikatem
                return JsonResponse({'success': False, 'message': 'Event jest już wyłączony.'})
            else:
                # Wyłącz zapisy na bal
                active_event.aktywny = False
                active_event.save()

                # Ustaw komunikat sukcesu w sesji
                messages.success(request, 'Zapisy na event zostały wyłączone.')

                # Przekieruj na stronę panel-index
                return redirect('panel:panel-index')

        return JsonResponse({'success': False, 'message': 'Brak aktywnego eventu.'})

    
def wypelnij_dane_bal(row_data, field_name, col_num, ws, row_num):
    col_letter = get_column_letter(col_num)

    # Mapowanie nazw pól modelu na odpowiadające im kolumny
    field_mapping = {
        'Lp': 'lp',
        'Login': 'login',
        'Imie': 'imie',
        'Nazwisko': 'nazwisko',
        'Data Utworzenia': 'data_utworzenia',
    }

    field_name_bal = field_mapping.get(field_name, field_name.lower())

    cell_value = row_data.get(field_name_bal, '')

    # Dodaj konwersję daty do łańcucha znakowego
    if isinstance(cell_value, datetime):
        cell_value = cell_value.strftime("%d-%m-%Y %H:%M")

    ws.cell(row=row_num, column=col_num, value=cell_value)
    print(f"Bal - row_data: {row_data}")
    print(f"field_name: {field_name_bal}, cell_value: {cell_value}, row_data: {row_data}")



def wypelnij_dane_piknik(row_data, field_name, col_num, ws, row_num):
    col_letter = get_column_letter(col_num)

    # Mapowanie nazw pól modelu na odpowiadające im kolumny
    field_mapping = {
        'Lp': 'lp',
        'Badge': 'identyfikator',
        'Login': 'login',
        'Imie': 'imie',
        'Nazwisko': 'nazwisko',
        'Osoba Towarzyszaca': 'osoba_towarzyszaca',
        'Liczba Dzieci': 'liczba_dzieci',
        'Wiek dziecka 1': 'wiek_dziecka_1',
        'Wiek dziecka 2': 'wiek_dziecka_2',
        'Wiek dziecka 3': 'wiek_dziecka_3',
        'Wiek dziecka 4': 'wiek_dziecka_4',
        'Wiek dziecka 5': 'wiek_dziecka_5',
        'Wiek dziecka 6': 'wiek_dziecka_6',
        'Wiek dziecka 7': 'wiek_dziecka_7',
        'Transport własny': 'transport_wlasny',
        'Przystanek': 'przystanek',
        'Regulamin pikniku': 'zaakceptowane_regulamin',
        'Data Utworzenia': 'data_utworzenia',
        
    }

    field_name = field_mapping.get(field_name, field_name.lower())

    cell_value = row_data.get(field_name, '')

    # Dodaj obsługę obiektów Przystanek
    if isinstance(cell_value, Przystanek):
        cell_value = str(cell_value)

    # Dodaj konwersję daty do łańcucha znakowego
    elif isinstance(cell_value, datetime):
        cell_value = cell_value.strftime("%d-%m-%Y %H:%M")

    # Jeżeli pole to osoba_towarzyszaca, zamień wartość na 'Tak' lub 'Nie'
    elif field_name == 'osoba_towarzyszaca':
        cell_value = 'Tak' if cell_value else 'Nie'

    # Jeżeli pole to Transport wlasny, zamień wartość na 'Tak' lub 'Nie'
    elif field_name == 'transport_wlasny':
        cell_value = 'Tak' if cell_value else 'Nie'

    # Jeżeli pole to akceptacja regulaminu, zamień wartość na 'Tak' lub 'Nie'
    elif field_name == 'zaakceptowane_regulamin':
        cell_value = 'Tak' if cell_value else 'Nie'

    ws.cell(row=row_num, column=col_num, value=cell_value)
    print(f"field_name: {field_name}, cell_value: {cell_value}, row_data: {row_data}")

# Main function
def generate_excel_report(request):
    active_event = ActiveEvent.objects.filter(aktywny=True).first()

    # Inicjalizuj dane jako puste
    data = []

    # Sprawdź, czy aktywny event jest Bal czy Piknik
    if active_event and active_event.rodzaj_eventu == 'bal':
        # Pobierz dane z modelu Bal
        bal_data = Bal.objects.filter(event=active_event, is_registred=True)
        # Przetwórz dane do formatu, który będzie używany w szablonie
        data = [{'lp': idx + 1, 'login': bal.login, 'imie': bal.imie, 'nazwisko': bal.nazwisko, 'data_utworzenia': bal.data_utworzenia} for idx, bal in enumerate(bal_data)]

    elif active_event and active_event.rodzaj_eventu == 'piknik':
        # Pobierz dane z modelu Piknik
        piknik_data = Piknik.objects.filter(event=active_event, is_registred=True)
        # Przetwórz dane do formatu, który będzie używany w szablonie
        data = [{'lp': idx + 1, 'identyfikator': piknik.identyfikator, 'login': piknik.login, 'imie': piknik.imie, 'nazwisko': piknik.nazwisko,
                 'osoba_towarzyszaca': piknik.osoba_towarzyszaca, 'liczba_dzieci': piknik.liczba_dzieci, 'wiek_dziecka_1': piknik.wiek_dziecka_1, 'wiek_dziecka_2': piknik.wiek_dziecka_2, 'wiek_dziecka_3': piknik.wiek_dziecka_3, 'wiek_dziecka_4': piknik.wiek_dziecka_4, 'wiek_dziecka_5': piknik.wiek_dziecka_5, 'wiek_dziecka_6': piknik.wiek_dziecka_6, 'wiek_dziecka_7': piknik.wiek_dziecka_7,
                 'transport_wlasny': piknik.transport_wlasny, 'przystanek': piknik.przystanek, 'zaakceptowane_regulamin': piknik.zaakceptowane_regulamin, 'data_utworzenia': piknik.data_utworzenia} for idx, piknik in enumerate(piknik_data)]

    # Tworzymy nowy arkusz Excela
    wb = openpyxl.Workbook()
    ws = wb.active

    # Nagłówki
    headers = ['Lp', 'Identyfikator', 'Login', 'Imie', 'Nazwisko']
    if active_event and active_event.rodzaj_eventu == 'piknik':
        headers += ['Osoba Towarzyszaca', 'Liczba Dzieci', 'Wiek dziecka 1', 'Wiek dziecka 2', 'Wiek dziecka 3', 'Wiek dziecka 4', 'Wiek dziecka 5', 'Wiek dziecka 6', 'Wiek dziecka 7', 'Transport własny', 'Przystanek', 'Regulamin pikniku', 'Data Utworzenia']
    else:
        headers += ['Data Utworzenia']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)

    # Wypełniamy danymi
    for row_num, row_data in enumerate(data, 2):
        for col_num, field_name in enumerate(headers, 1):
            if active_event and active_event.rodzaj_eventu == 'bal':
                wypelnij_dane_bal(row_data, field_name, col_num, ws, row_num)
            elif active_event and active_event.rodzaj_eventu == 'piknik':
                wypelnij_dane_piknik(row_data, field_name, col_num, ws, row_num)

    # Dodaj datę i godzinę do nazwy pliku
    current_datetime = datetime.now().strftime("%Y%m%d-%H%M%S")
    event_type = active_event.rodzaj_eventu
    filename = f"{event_type}_{current_datetime}.xlsx"

    # Tworzymy odpowiedź HTTP z plikiem Excela
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    wb.close()

    return response

#---------------------------------------- EXCEL - WYPISANI --------------------------------#

#from django.db.models import F

# Funkcja do wypełniania danych dla listy wypisanych
def wypelnij_dane_wypisanych_bal(row_data, field_name, col_num, ws, row_num):
    col_letter = get_column_letter(col_num)

    # Mapowanie nazw pól modelu na odpowiadające im kolumny
    field_mapping = {
        'Lp': 'lp',
        'Login': 'login',
        'Imie': 'imie',
        'Nazwisko': 'nazwisko',
        #'Data Utworzenia': 'data_utworzenia',
        'Data Modyfikacji': 'data_modyfikacji',  # Dodane pole daty modyfikacji
    }

    field_name_wypisani = field_mapping.get(field_name, field_name.lower())

    cell_value = row_data.get(field_name_wypisani, '')

    # Dodaj konwersję daty do łańcucha znakowego
    if isinstance(cell_value, datetime):
        cell_value = cell_value.strftime("%d-%m-%Y %H:%M")

    ws.cell(row=row_num, column=col_num, value=cell_value)
    print(f"Wypisani - row_data: {row_data}")
    print(f"field_name: {field_name_wypisani}, cell_value: {cell_value}, row_data: {row_data}")

# Funkcja do wypełniania danych dla listy wypisanych
def wypelnij_dane_wypisanych_piknik(row_data, field_name, col_num, ws, row_num):
    col_letter = get_column_letter(col_num)

    # Mapowanie nazw pól modelu na odpowiadające im kolumny
    field_mapping = {
        'Lp': 'lp',
        'Badge': 'identyfikator',
        'Login': 'login',
        'Imie': 'imie',
        'Nazwisko': 'nazwisko',
        'Osoba Towarzyszaca': 'osoba_towarzyszaca',
        'Liczba Dzieci': 'liczba_dzieci',
        'Wiek dziecka 1': 'wiek_dziecka_1',
        'Wiek dziecka 2': 'wiek_dziecka_2',
        'Wiek dziecka 3': 'wiek_dziecka_3',
        'Wiek dziecka 4': 'wiek_dziecka_4',
        'Wiek dziecka 5': 'wiek_dziecka_5',
        'Wiek dziecka 6': 'wiek_dziecka_6',
        'Wiek dziecka 7': 'wiek_dziecka_7',
        'Transport własny': 'transport_wlasny',
        'Przystanek': 'przystanek',
        'Regulamin pikniku': 'zaakceptowane_regulamin',
        #'Data Utworzenia': 'data_utworzenia',
        'Data Modyfikacji': 'data_modyfikacji',  # Dodane pole daty modyfikacji
    }

    field_name_wypisani = field_mapping.get(field_name, field_name.lower())

    cell_value = row_data.get(field_name_wypisani, '')

    # Dodaj obsługę obiektów Przystanek
    if isinstance(cell_value, Przystanek):
        cell_value = str(cell_value)

    # Dodaj konwersję daty do łańcucha znakowego
    elif isinstance(cell_value, datetime):
        cell_value = cell_value.strftime("%d-%m-%Y %H:%M")

    # Jeżeli pole to osoba_towarzyszaca, zamień wartość na 'Tak' lub 'Nie'
    elif field_name_wypisani == 'osoba_towarzyszaca':
        cell_value = 'Tak' if cell_value else 'Nie'

    # Jeżeli pole to osoba_towarzyszaca, zamień wartość na 'Tak' lub 'Nie'
    elif field_name_wypisani == 'transport_wlasny':
        cell_value = 'Tak' if cell_value else 'Nie'
    
    # Jeżeli pole to akceptacja regulaminu, zamień wartość na 'Tak' lub 'Nie'
    elif field_name_wypisani == 'zaakceptowane_regulamin':
        cell_value = 'Tak' if cell_value else 'Nie'

    ws.cell(row=row_num, column=col_num, value=cell_value)
    print(f"field_name: {field_name_wypisani}, cell_value: {cell_value}, row_data: {row_data}")

# Funkcja do generowania raportu dla wypisanych
def generate_excel_report_wypisani(request):
    active_event = ActiveEvent.objects.filter(aktywny=True).first()

    # Inicjalizuj dane jako puste
    data = []

    # Sprawdź, czy aktywny event jest Bal czy Piknik
    if active_event and active_event.rodzaj_eventu == 'bal':
        # Pobierz dane z modelu Bal dla wypisanych
        bal_data = Bal.objects.filter(event=active_event, is_registred=False)
        # Przetwórz dane do formatu, który będzie używany w szablonie
        data = [{'lp': idx + 1, 'login': bal.login, 'imie': bal.imie, 'nazwisko': bal.nazwisko,
                 'data_utworzenia': bal.data_utworzenia, 'data_modyfikacji': bal.data_modyfikacji} for idx, bal in enumerate(bal_data)]

    elif active_event and active_event.rodzaj_eventu == 'piknik':
        # Pobierz dane z modelu Piknik dla wypisanych
        piknik_data = Piknik.objects.filter(event=active_event, is_registred=False)
        # Przetwórz dane do formatu, który będzie używany w szablonie
        data = [{'lp': idx + 1, 'identyfikator': piknik.identyfikator, 'login': piknik.login, 'imie': piknik.imie, 'nazwisko': piknik.nazwisko,
                 'osoba_towarzyszaca': piknik.osoba_towarzyszaca, 'liczba_dzieci': piknik.liczba_dzieci, 'wiek_dziecka_1': piknik.wiek_dziecka_1, 'wiek_dziecka_2': piknik.wiek_dziecka_2, 'wiek_dziecka_3': piknik.wiek_dziecka_3, 'wiek_dziecka_4': piknik.wiek_dziecka_4, 'wiek_dziecka_5': piknik.wiek_dziecka_5, 'wiek_dziecka_6': piknik.wiek_dziecka_6, 'wiek_dziecka_7': piknik.wiek_dziecka_7,
                 'transport_wlasny': piknik.transport_wlasny, 'przystanek': piknik.przystanek, 'zaakceptowane_regulamin': piknik.zaakceptowane_regulamin, 'data_utworzenia': piknik.data_utworzenia,
                 'data_modyfikacji': piknik.data_modyfikacji} for idx, piknik in enumerate(piknik_data)]

    # Tworzymy nowy arkusz Excela
    wb = openpyxl.Workbook()
    ws = wb.active

    # Nagłówki
    headers = ['Lp', 'Badge', 'Login', 'Imie', 'Nazwisko']
    if active_event and active_event.rodzaj_eventu == 'piknik':
        headers += ['Osoba Towarzyszaca', 'Liczba Dzieci', 'Wiek dziecka 1', 'Wiek dziecka 2', 'Wiek dziecka 3', 'Wiek dziecka 4', 'Wiek dziecka 5', 'Wiek dziecka 6', 'Wiek dziecka 7', 'Transport własny', 'Przystanek', 'Regulamin pikniku', 'Data Modyfikacji']
    else:
        #headers += ['Data Utworzenia', 'Data Modyfikacji']
        headers += ['Data Modyfikacji']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)

    # Wypełniamy danymi
    for row_num, row_data in enumerate(data, 2):
        for col_num, field_name in enumerate(headers, 1):
            if active_event and active_event.rodzaj_eventu == 'bal':
                wypelnij_dane_wypisanych_bal(row_data, field_name, col_num, ws, row_num)
            elif active_event and active_event.rodzaj_eventu == 'piknik':
                wypelnij_dane_wypisanych_piknik(row_data, field_name, col_num, ws, row_num)

    # Dodaj datę i godzinę do nazwy pliku
    current_datetime = datetime.now().strftime("%Y%m%d-%H%M%S")
    event_type = active_event.rodzaj_eventu
    filename = f"{event_type}_wypisani_{current_datetime}.xlsx"

    # Tworzymy odpowiedź HTTP z plikiem Excela
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    wb.close()

    return response


#---------------- Raport zapisanych na piknik - transport wlasny ------------------------#

def excel_report_piknik_transport(request):
    active_event = ActiveEvent.objects.filter(aktywny=True).first()

    # Inicjalizuj dane jako puste
    data = []

    # Sprawdź, czy aktywny event jest Piknikiem
    if active_event and active_event.rodzaj_eventu == 'piknik':
        # Pobierz dane z modelu Piknik
        piknik_data = Piknik.objects.filter(event=active_event, is_registred=True, transport_wlasny=True)
        # Przetwórz dane do formatu, który będzie używany w szablonie
        data = [{'lp': idx + 1, 'identyfikator': piknik.identyfikator, 'login': piknik.login, 'imie': piknik.imie, 'nazwisko': piknik.nazwisko,
                 'osoba_towarzyszaca': piknik.osoba_towarzyszaca, 'liczba_dzieci': piknik.liczba_dzieci, 'wiek_dziecka_1': piknik.wiek_dziecka_1, 'wiek_dziecka_2': piknik.wiek_dziecka_2, 'wiek_dziecka_3': piknik.wiek_dziecka_3, 'wiek_dziecka_4': piknik.wiek_dziecka_4, 'wiek_dziecka_5': piknik.wiek_dziecka_5, 'wiek_dziecka_6': piknik.wiek_dziecka_6, 'wiek_dziecka_7': piknik.wiek_dziecka_7,
                 'transport_wlasny': piknik.transport_wlasny, 'zaakceptowane_regulamin': piknik.zaakceptowane_regulamin, 'data_utworzenia': piknik.data_utworzenia} for idx, piknik in enumerate(piknik_data)]

    # Tworzymy nowy arkusz Excela
    wb = openpyxl.Workbook()
    ws = wb.active

    # Nagłówki
    headers = ['Lp', 'Badge', 'Login', 'Imie', 'Nazwisko', 'Osoba Towarzyszaca', 'Liczba Dzieci', 'Wiek dziecka 1', 'Wiek dziecka 2', 'Wiek dziecka 3', 'Wiek dziecka 4', 'Wiek dziecka 5', 'Wiek dziecka 6', 'Wiek dziecka 7', 'Transport własny',  'Regulamin pikniku', 'Data Utworzenia']
    
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)

    # Wypełniamy danymi
    for row_num, row_data in enumerate(data, 2):
        for col_num, field_name in enumerate(headers, 1):
            wypelnij_dane_piknik(row_data, field_name, col_num, ws, row_num)

    # Dodaj datę i godzinę do nazwy pliku
    current_datetime = datetime.now().strftime("%Y%m%d-%H%M%S")
    event_type = active_event.rodzaj_eventu
    filename = f"{event_type}_wlasny_{current_datetime}.xlsx"

    # Tworzymy odpowiedź HTTP z plikiem Excela
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    wb.close()

    return response

#---------------- Raport zapisanych na piknik - transport amazonski ------------------------#

def excel_report_piknik_przystanek(request):
    active_event = ActiveEvent.objects.filter(aktywny=True).first()

    # Inicjalizuj dane jako puste
    data = []

    # Sprawdź, czy aktywny event jest Piknikiem
    if active_event and active_event.rodzaj_eventu == 'piknik':
        # Pobierz dane z modelu Piknik
        piknik_data = Piknik.objects.filter(event=active_event, is_registred=True, transport_wlasny=False)
        # Przetwórz dane do formatu, który będzie używany w szablonie
        data = [{'lp': idx + 1, 'identyfikator': piknik.identyfikator, 'login': piknik.login, 'imie': piknik.imie, 'nazwisko': piknik.nazwisko,
                 'osoba_towarzyszaca': piknik.osoba_towarzyszaca, 'liczba_dzieci': piknik.liczba_dzieci, 'wiek_dziecka_1': piknik.wiek_dziecka_1, 'wiek_dziecka_2': piknik.wiek_dziecka_2, 'wiek_dziecka_3': piknik.wiek_dziecka_3, 'wiek_dziecka_4': piknik.wiek_dziecka_4, 'wiek_dziecka_5': piknik.wiek_dziecka_5, 'wiek_dziecka_6': piknik.wiek_dziecka_6, 'wiek_dziecka_7': piknik.wiek_dziecka_7,
                 'przystanek': piknik.przystanek, 'zaakceptowane_regulamin': piknik.zaakceptowane_regulamin, 'data_utworzenia': piknik.data_utworzenia} for idx, piknik in enumerate(piknik_data)]

    # Tworzymy nowy arkusz Excela
    wb = openpyxl.Workbook()
    ws = wb.active

    # Nagłówki
    headers = ['Lp', 'Badge', 'Login', 'Imie', 'Nazwisko', 'Osoba Towarzyszaca', 'Liczba Dzieci', 'Wiek dziecka 1', 'Wiek dziecka 2', 'Wiek dziecka 3', 'Wiek dziecka 4', 'Wiek dziecka 5', 'Wiek dziecka 6', 'Wiek dziecka 7', 'Przystanek',  'Regulamin pikniku', 'Data Utworzenia']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header
        ws[f'{col_letter}1'].font = Font(bold=True)

    # Wypełniamy danymi
    for row_num, row_data in enumerate(data, 2):
        for col_num, field_name in enumerate(headers, 1):
            wypelnij_dane_piknik(row_data, field_name, col_num, ws, row_num)

    # Dodaj datę i godzinę do nazwy pliku
    current_datetime = datetime.now().strftime("%Y%m%d-%H%M%S")
    event_type = active_event.rodzaj_eventu
    filename = f"{event_type}_amazon_{current_datetime}.xlsx"

    # Tworzymy odpowiedź HTTP z plikiem Excela
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    wb.close()

    return response